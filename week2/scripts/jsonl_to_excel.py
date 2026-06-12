#!/usr/bin/env python3
"""Convert Week 2 JSONL files to Excel workbook with summary sheets."""
import json, os
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

JSONL_DIR = Path(__file__).resolve().parent.parent / "outputs" / "week2_jsonl"
EXCEL_DIR = Path(__file__).resolve().parent.parent / "outputs" / "week2_excel"
os.makedirs(EXCEL_DIR, exist_ok=True)

OUT_XLSX = EXCEL_DIR / "week2_extraction.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

COMPANY_ORDER = [
    ("603418", "友升股份"), ("001282", "三联锻造"), ("301581", "黄山谷捷"),
    ("301563", "云汉芯城"), ("688758", "赛分科技"), ("688775", "影石创新"),
    ("920100", "三协电机"), ("920116", "星图测控"),
]

SUB_FIELDS = [
    "record_type", "stock_code", "company_name", "pdf_page",
    "subscription_date", "batch_label", "subscriber_name",
    "subscription_shares_wan", "subscription_amount_wan", "subscription_price_yuan",
    "evidence_text", "notes"
]
EQ_FIELDS = [
    "record_type", "stock_code", "company_name", "pdf_page",
    "time_point", "equity_structure_scope",
    "total_shares_wan", "total_capital_wan",
    "shareholder_name", "shares_held_wan", "capital_contribution_wan",
    "shareholding_ratio", "evidence_text", "notes"
]

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# Summary sheet
ws_summary = wb.create_sheet("Summary", 0)
summary_headers = ["公司代码", "公司简称", "Subscription Flow", "Equity Snapshot", "合计"]
for c, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER

summary_row = 2
grand_sub, grand_eq = 0, 0

for code, name in COMPANY_ORDER:
    fp = JSONL_DIR / f"{code}_{name}.jsonl"
    sub_count, eq_count = 0, 0
    sub_records, eq_records = [], []

    if fp.exists():
        with open(fp, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except:
                    continue
                if obj.get("record_type") == "subscription_flow":
                    sub_records.append(obj)
                    sub_count += 1
                elif obj.get("record_type") == "equity_snapshot":
                    eq_records.append(obj)
                    eq_count += 1

    grand_sub += sub_count
    grand_eq += eq_count
    ws_summary.cell(row=summary_row, column=1, value=code).font = DATA_FONT
    ws_summary.cell(row=summary_row, column=2, value=name).font = DATA_FONT
    ws_summary.cell(row=summary_row, column=3, value=sub_count).font = DATA_FONT
    ws_summary.cell(row=summary_row, column=4, value=eq_count).font = DATA_FONT
    ws_summary.cell(row=summary_row, column=5, value=sub_count + eq_count).font = DATA_FONT
    for c in range(1, 6):
        ws_summary.cell(row=summary_row, column=c).border = THIN_BORDER
        ws_summary.cell(row=summary_row, column=c).alignment = Alignment(horizontal="center")
    summary_row += 1

# Total row
ws_summary.cell(row=summary_row, column=1, value="合计").font = Font(name="微软雅黑", size=10, bold=True)
ws_summary.cell(row=summary_row, column=2, value="").font = DATA_FONT
ws_summary.cell(row=summary_row, column=3, value=grand_sub).font = Font(name="微软雅黑", size=10, bold=True)
ws_summary.cell(row=summary_row, column=4, value=grand_eq).font = Font(name="微软雅黑", size=10, bold=True)
ws_summary.cell(row=summary_row, column=5, value=grand_sub + grand_eq).font = Font(name="微软雅黑", size=10, bold=True)
for c in range(1, 6):
    ws_summary.cell(row=summary_row, column=c).border = THIN_BORDER
    ws_summary.cell(row=summary_row, column=c).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=summary_row, column=c).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ws_summary.column_dimensions["A"].width = 14
ws_summary.column_dimensions["B"].width = 16
ws_summary.column_dimensions["C"].width = 20
ws_summary.column_dimensions["D"].width = 20
ws_summary.column_dimensions["E"].width = 12

def write_sheet(ws, headers, records, sheet_name):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for r, rec in enumerate(records, 2):
        for c, field in enumerate(headers, 1):
            val = rec.get(field, "")
            if val is None:
                val = ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

    # Auto-width
    col_widths = {
        "subscription_flow": {
            "record_type": 16, "stock_code": 12, "company_name": 14, "pdf_page": 10,
            "subscription_date": 16, "batch_label": 10, "subscriber_name": 22,
            "subscription_shares_wan": 18, "subscription_amount_wan": 18,
            "subscription_price_yuan": 16, "evidence_text": 60, "notes": 20
        },
        "equity_snapshot": {
            "record_type": 16, "stock_code": 12, "company_name": 14, "pdf_page": 10,
            "time_point": 14, "equity_structure_scope": 20,
            "total_shares_wan": 16, "total_capital_wan": 16,
            "shareholder_name": 22, "shares_held_wan": 16,
            "capital_contribution_wan": 18, "shareholding_ratio": 14,
            "evidence_text": 60, "notes": 20
        }
    }

    widths = col_widths.get(ws.title.split("_")[0], {})
    for c, h in enumerate(headers, 1):
        if h in widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[h]

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

# Create per-company sheets
SUB_KEY = "subscription_flow"
EQ_KEY = "equity_snapshot"

for code, name in COMPANY_ORDER:
    fp = JSONL_DIR / f"{code}_{name}.jsonl"
    if not fp.exists():
        continue

    sub_recs, eq_recs = [], []
    with open(fp, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except:
                continue
            if obj.get("record_type") == SUB_KEY:
                sub_recs.append(obj)
            elif obj.get("record_type") == EQ_KEY:
                eq_recs.append(obj)

    if sub_recs:
        ws = wb.create_sheet(f"{code}_sub")
        write_sheet(ws, SUB_FIELDS, sub_recs, f"{code}_sub")

    if eq_recs:
        ws = wb.create_sheet(f"{code}_eq")
        write_sheet(ws, EQ_FIELDS, eq_recs, f"{code}_eq")

wb.save(OUT_XLSX)
print(f"Excel written: {OUT_XLSX}")
print(f"Summary: {grand_sub} subscription_flow + {grand_eq} equity_snapshot = {grand_sub + grand_eq} total records")
